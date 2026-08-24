# -*- coding: utf-8 -*-
"""
연료·탄소 가격 무료 수집기 (TTF / EUA / API2계열 석탄) — GitHub Actions 전용.

■ 왜 entsoe-harvester 가 아니라 이 레포인가 (반복 질문 방지)
  용도는 WA(폴란드)에서 출발했지만 **데이터가 폴란드용이 아니다.**
  World Bank 파일 한 번 다운로드로 5계열이 나온다 — TTF(유럽)·Henry Hub(미국)·
  LNG(일본)·석탄 AU·석탄 ZA. 폴란드용은 그중 하나다. 특히 Henry Hub 는 같은 레포의
  fetch_steo.py 산출물과 **교차검증**해야 하므로 여기 있어야 대조가 된다.
  또 build_panel.py 가 이 레포에 있어서, 같은 레포 파일은 디스크에서 읽는다 —
  entsoe-harvester 에 두면 크로스레포 raw 배관을 새로 깔아야 한다.
  결정적으로 entsoe-harvester 잡은 스텝 실패 격리가 없어서, 여기 파싱 실패가
  **폴란드 도매가 수집까지 같이 죽인다.**
  (이 레포 이름은 초기 잔재다. 실제로는 EIA·ECB FX·PJM·MISO·STEO 를 담은 수집 허브다.
   리네임하면 파일럿의 GH_RAW 와 build_panel 의 레포 상수가 깨지므로 하지 않는다.)

■ 배경
  WA(폴란드) SRMC 경로의 EUA·TTF·API2 는 지금까지 **수동 입력**이었다.
  method.md 판정 "무료 공식 API 없음(ICE 계열)" 은 **forward 곡선**에 대한 것이다.
  과거 월별 현물은 사정이 다르다 — World Bank·IMF 가 월평균을 무료로 공개한다.
  파일럿에 부족한 것은 둘로 나뉜다.
    (A) 과거 월별 현물   → SRMC 계수(SRMC_BETA 등 v7 이식 상수)를 PL 실측으로 재추정·검증
    (B) 미래 forward 곡선 → FUEL_FWD 채우기
  (A) 없이 (B)만 구해도 계수를 검증할 수 없다. 그래서 (A)를 먼저 자동화한다.

■ 원칙
  1. URL 을 추측해서 박지 않는다. 후보를 순회하고 **Actions 로그가 판정**한다(IESO 백필과 동일).
  2. 파서를 쓸 수 없는 후보는 상태코드·본문머리만 찍는 **프로브**로 둔다.
  3. 실패는 조용히 넘기되 fuel_probe.json 에 전부 기록한다. 종료코드는 항상 0.
  4. 값을 만들어내지 않는다. 못 받은 계열은 출력 자체를 하지 않는다.
  5. 계열 0개면 기존 출력파일을 덮어쓰지 않는다(멀쩡한 과거 수집분 보호).

■ 산출
  fuel_monthly.json / .csv   파서 성공 계열만. steo_forward_monthly.json 과 동일 스키마
  fuel_probe.json            후보별 판정표 (다음 단계 결정 근거)

■ EUA 전용 블록 (2026-08-24)
  EUA 는 폴란드 도매 공정가의 51% 를 차지한다(아래 T1-EUA 주석의 계산 참조).
  1차 실행에서 유일하게 실패한 항목이라 별도 경로 4가지로 다시 판다.

■ 사용
  python fetch_fuel.py              수집 + 프로브
  python fetch_fuel.py --discover   프로브만 (출력파일 미갱신)

■ 의존
  requests, openpyxl   (collect.yml 의 pip 줄에 openpyxl 추가 필요)
"""
import os, sys, json, csv, re, io, time
import requests

UA = {"User-Agent": "Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 "
                    "(KHTML, like Gecko) Chrome/126.0 Safari/537.36"}
TIMEOUT   = 60
OUT_JSON  = "fuel_monthly.json"
OUT_CSV   = "fuel_monthly.csv"
OUT_PROBE = "fuel_probe.json"
DISCOVER  = "--discover" in sys.argv

PROBE  = []     # [{id, tier, url, http, ok, note}]
SERIES = {}     # {key: {label, unit, source, points:[{period,value}]}}


def log(*a):
    print(*a, flush=True)


def rec(pid, tier, url, http=None, ok=False, note=""):
    PROBE.append({"id": pid, "tier": tier, "url": url,
                  "http": http, "ok": bool(ok), "note": note[:400]})
    log("[%s] %-22s tier%s http=%s %s" %
        ("OK " if ok else "NG ", pid, tier, http, note[:160]))


def get(url, **kw):
    kw.setdefault("headers", UA)
    kw.setdefault("timeout", TIMEOUT)
    return requests.get(url, **kw)


def put(key, label, unit, source, pairs):
    """pairs = [(YYYY-MM, float)] — 빈 계열은 등록하지 않는다."""
    pts = [{"period": p, "value": v} for p, v in sorted(pairs) if v is not None]
    if not pts:
        return 0
    SERIES[key] = {"label": label, "unit": unit, "source": source, "points": pts}
    return len(pts)


# ══════════════════════════════════════════════════════════════════
#  xlsx 관용 파서 — 시트 구조가 바뀌어도 날짜열·헤더행을 스스로 찾는다
#    ① 날짜열 = 'YYYYMnn' 매치가 가장 많은 열
#    ② 헤더    = 첫 날짜행 위 20행을 **열 단위로 이어붙인 문자열**(단일 헤더행 가정 금지)
#  World Bank·IMF 둘 다 상단에 제목·단위행이 섞여 있어 고정 인덱스로 읽으면 깨진다.
# ══════════════════════════════════════════════════════════════════
YM_RE = re.compile(r"^\s*(\d{4})\s*[Mm]\s*(\d{1,2})\s*$")


def xlsx_grid(blob, sheet_hint="month", max_rows=2200):
    from openpyxl import load_workbook
    wb = load_workbook(io.BytesIO(blob), data_only=True, read_only=True)
    name = None
    for n in wb.sheetnames:
        if sheet_hint.lower() in n.lower():
            name = n
            break
    name = name or wb.sheetnames[0]
    ws = wb[name]
    grid = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i >= max_rows:
            break
        grid.append(list(row))
    wb.close()
    return name, grid


def xlsx_pick(grid, targets):
    """targets = {key: (헤더정규식, 라벨)} -> ({key: [(ym, val)]}, headerRow, dateCol)"""
    ncol = max((len(r) for r in grid), default=0)
    if not ncol:
        return {}, -1, -1
    hit = [0] * ncol
    for r in grid:
        for c in range(min(ncol, len(r))):
            if r[c] is not None and YM_RE.match(str(r[c])):
                hit[c] += 1
    if not any(hit):
        return {}, -1, -1
    dcol = hit.index(max(hit))
    first = next(i for i, r in enumerate(grid)
                 if len(r) > dcol and r[dcol] is not None and YM_RE.match(str(r[dcol])))

    # 헤더는 **한 행이 아니다.** World Bank 는 품목명행과 단위행이 분리돼 있고
    # IMF 는 품목명·코드·단위가 3행으로 쪼개져 있다. 한 행만 고르면 단위행("($/mmbtu)")을
    # 골라 전부 미스한다. → 데이터 시작행 위쪽 전체를 **열 단위로 이어붙여** 매칭한다.
    pre0 = max(0, first - 20)
    header = []
    for c in range(ncol):
        buf = []
        for i in range(pre0, first):
            if len(grid[i]) > c and isinstance(grid[i][c], str) and grid[i][c].strip():
                buf.append(grid[i][c].strip())
        header.append(" ".join(buf))
    hrow = "%d~%d(concat)" % (pre0, first - 1)

    out = {}
    for key, (rx, _label) in targets.items():
        col = None
        for c, h in enumerate(header):
            if h and re.search(rx, h, re.I):
                col = c
                break
        if col is None:
            continue
        pairs = []
        for r in grid[first:]:
            if len(r) <= max(col, dcol) or r[dcol] is None:
                continue
            m = YM_RE.match(str(r[dcol]))
            if not m:
                continue
            ym = "%04d-%02d" % (int(m.group(1)), int(m.group(2)))
            v = r[col]
            try:
                v = float(str(v).replace(",", "")) if v not in (None, "", "..", "n/a") else None
            except (TypeError, ValueError):
                v = None
            if v is not None:
                pairs.append((ym, v))
        if pairs:
            out[key] = pairs
    return out, hrow, dcol


# ══════════════════════════════════════════════════════════════════
#  T1  World Bank Pink Sheet (월별·무키) — 링크를 페이지에서 발견해서 쓴다
#      해시경로가 릴리스마다 바뀌므로 URL 을 코드에 고정하면 언젠가 죽는다.
# ══════════════════════════════════════════════════════════════════
WB_PAGE = "https://www.worldbank.org/en/research/commodity-markets"
WB_FALLBACK = [
    # 페이지 정규식이 실패할 때만 시도하는 과거 관측 경로. 살아있다고 단정하지 않는다.
    "https://thedocs.worldbank.org/en/doc/18675f1d1639c7a34d463f59263ba0a2-0050012025/related/CMO-Historical-Data-Monthly.xlsx",
    "https://thedocs.worldbank.org/en/doc/5d903e848db1d1b83e0ec8f744e55570-0350012021/related/CMO-Historical-Data-Monthly.xlsx",
]
WB_TARGETS = {
    "TTF_EU":  (r"natural\s*gas.*europe",       "Natural gas, Europe (TTF 연동)"),
    "HH_US":   (r"natural\s*gas.*\bu\.?s\.?\b", "Natural gas, US (Henry Hub)"),
    "LNG_JP":  (r"liquefied|\blng\b",           "LNG, Japan"),
    "COAL_ZA": (r"coal.*south\s*afric",         "Coal, South Africa (API4계열)"),
    "COAL_AU": (r"coal.*austral",               "Coal, Australia (Newcastle)"),
}


def src_worldbank():
    urls = []
    try:
        r = get(WB_PAGE)
        found = re.findall(r'https://[^"\'\)\s]*CMO-Historical-Data-Monthly\.xlsx', r.text, re.I)
        urls = list(dict.fromkeys(found))
        rec("WB_page", 1, WB_PAGE, r.status_code, bool(urls),
            ("xlsx링크 %d건 발견" % len(urls)) if urls else "페이지 정규식 실패 -> 폴백 시도")
    except Exception as e:
        rec("WB_page", 1, WB_PAGE, None, False, "EXC %s" % e)
    urls = urls + WB_FALLBACK

    for u in urls:
        try:
            r = get(u)
            if r.status_code != 200 or len(r.content) < 20000:
                rec("WB_xlsx", 1, u, r.status_code, False, "len=%d" % len(r.content))
                continue
            sheet, grid = xlsx_grid(r.content, "month")
            got, hrow, dcol = xlsx_pick(grid, WB_TARGETS)
            n = 0
            for k, pairs in got.items():
                n += put("WB_" + k, WB_TARGETS[k][1], "USD(원문단위)",
                         "World Bank Pink Sheet monthly", pairs)
            rec("WB_xlsx", 1, u, r.status_code, n > 0,
                "sheet=%s headerRow=%s dateCol=%s 계열=%s pts=%d" % (sheet, hrow, dcol, list(got), n))
            if n:
                return
        except Exception as e:
            rec("WB_xlsx", 1, u, None, False, "EXC %s" % e)


# ══════════════════════════════════════════════════════════════════
#  T1  IMF Primary Commodity Prices (월별·무키)
# ══════════════════════════════════════════════════════════════════
IMF_URLS = [
    "https://www.imf.org/-/media/Files/Research/CommodityPrices/Monthly/external-data.ashx",
    "https://www.imf.org/-/media/Files/Research/CommodityPrices/Monthly/ExternalData.ashx",
]
IMF_TARGETS = {
    "NGAS_EU": (r"pngaseu|natural\s*gas.*(eu\b|europe|ttf)", "IMF Natural Gas, EU (TTF)"),
    "COAL_AU": (r"pcoalau|coal.*austral",                    "IMF Coal, Australia"),
    "COAL_SA": (r"pcoalsa|coal.*south",                      "IMF Coal, South Africa"),
}


def src_imf():
    for u in IMF_URLS:
        try:
            r = get(u)
            if r.status_code != 200 or len(r.content) < 20000:
                rec("IMF_xlsx", 1, u, r.status_code, False, "len=%d" % len(r.content))
                continue
            sheet, grid = xlsx_grid(r.content, "month")
            got, hrow, dcol = xlsx_pick(grid, IMF_TARGETS)
            n = 0
            for k, pairs in got.items():
                n += put("IMF_" + k, IMF_TARGETS[k][1], "USD(원문단위)",
                         "IMF Primary Commodity Prices", pairs)
            rec("IMF_xlsx", 1, u, r.status_code, n > 0,
                "sheet=%s headerRow=%s dateCol=%s 계열=%s pts=%d" % (sheet, hrow, dcol, list(got), n))
            if n:
                return
        except Exception as e:
            rec("IMF_xlsx", 1, u, None, False, "EXC %s" % e)


# ══════════════════════════════════════════════════════════════════
#  T1  Yahoo chart v8 (무키)
#      EUA 현물 티커는 무료로 안 열리지만, **EUA 선물을 실제로 보유하는 ETF**
#      (KEUA=EU전용, KRBN=글로벌·EUA 비중최대)는 열린다. 대리지표로 상관을 먼저 측정한다.
#      TTF=F / MTF=F 는 존재 여부 미확인 — 로그가 판정한다.
# ══════════════════════════════════════════════════════════════════
YH = [
    ("KEUA",  "KraneShares European Carbon ETF (EUA선물 보유)", "USD/주"),
    ("KRBN",  "KraneShares Global Carbon ETF (EUA 비중최대)",   "USD/주"),
    ("TTF=F", "Yahoo TTF 선물 후보 (존재여부 미확인)",            "?"),
    ("MTF=F", "Yahoo TTF 계열 후보 (미확인)",                    "?"),
    ("NG=F",  "Henry Hub 선물 (대조군·STEO 교차검증용)",          "USD/MMBtu"),
]


def src_yahoo():
    for sym, label, unit in YH:
        u = ("https://query1.finance.yahoo.com/v8/finance/chart/%s"
             "?range=10y&interval=1mo" % sym)
        try:
            r = get(u)
            if r.status_code != 200:
                rec("YH_" + sym, 1, u, r.status_code, False, r.text[:120])
                continue
            j = r.json()
            res = ((j.get("chart") or {}).get("result") or [None])[0]
            if not res:
                rec("YH_" + sym, 1, u, r.status_code, False,
                    str((j.get("chart") or {}).get("error"))[:160])
                continue
            ts = res.get("timestamp") or []
            ind = res.get("indicators") or {}
            vals = ((ind.get("adjclose") or [{}])[0].get("adjclose")
                    or (ind.get("quote") or [{}])[0].get("close") or [])
            pairs = []
            for t, v in zip(ts, vals):
                if v is None:
                    continue
                g = time.gmtime(t)
                pairs.append(("%04d-%02d" % (g.tm_year, g.tm_mon), float(v)))
            n = put("YH_" + sym.replace("=", ""), label, unit, "Yahoo Finance chart v8", pairs)
            rec("YH_" + sym, 1, u, r.status_code, n > 0,
                "pts=%d %s~%s" % (n, pairs[0][0] if pairs else "-", pairs[-1][0] if pairs else "-"))
        except Exception as e:
            rec("YH_" + sym, 1, u, None, False, "EXC %s" % e)


# ══════════════════════════════════════════════════════════════════
#  T2  Stooq 월별 CSV (무키) — 티커 존재여부가 불확실. 프로브 비용이 0 이라 전량 시도.
# ══════════════════════════════════════════════════════════════════
STOOQ = ["ttf.f", "co2.f", "eua.f", "atw.f", "coal.f", "ng.f", "ttfg.f", "carb.f"]


def src_stooq():
    for s in STOOQ:
        u = "https://stooq.com/q/d/l/?s=%s&i=m" % s
        try:
            r = get(u)
            body = r.text[:200]
            if r.status_code != 200 or not body.lower().startswith("date"):
                rec("STQ_" + s, 2, u, r.status_code, False, body.replace("\n", " ")[:120])
                continue
            pairs = []
            for row in csv.DictReader(io.StringIO(r.text)):
                d = (row.get("Date") or "")[:7]
                try:
                    v = float(row.get("Close"))
                except (TypeError, ValueError):
                    continue
                if re.match(r"^\d{4}-\d{2}$", d):
                    pairs.append((d, v))
            n = put("STQ_" + s.replace(".", "_"), "Stooq " + s, "?", "Stooq monthly CSV", pairs)
            rec("STQ_" + s, 2, u, r.status_code, n > 0,
                "pts=%d %s~%s" % (n, pairs[0][0] if pairs else "-", pairs[-1][0] if pairs else "-"))
        except Exception as e:
            rec("STQ_" + s, 2, u, None, False, "EXC %s" % e)


# ══════════════════════════════════════════════════════════════════
#  T1-EUA  배출권 전용 블록 (2026-08-24 추가)
#
#  ■ 왜 EUA 만 따로 파는가
#    파일럿 SRMC 계수로 계산하면 **EUA 가 폴란드 도매 공정가의 51% 를 차지한다.**
#      석탄 SRMC = 연료 34.3 + 탄소 65.6 + VOM 2 = 101.9 EUR/MWh  → 탄소 64.3%
#      가스 SRMC = 연료 78.2 + 탄소 29.2 + VOM 2 = 109.4          → 탄소 26.7%
#      혼합(석탄 β=0.6651)                        = 104.4          → **탄소 51.1%**
#      (EUA 79 EUR/t · API2 114 USD/t · TTF 43 EUR/MWh · coalEmis 0.83 · coalEff 0.41)
#    석탄 SRMC 에서 **탄소비가 연료비보다 크다**(65.6 vs 34.3). 석탄 비중이 높다는 건
#    곧 EUA 노출이 크다는 뜻이다.
#    민감도 — EUA €10 → 혼합 SRMC 6.76 EUR/MWh → **10.9 ₩/kWh**.
#    WA 차년도 예측 195.0 ₩/kWh 의 5.6% 로 **MAPE 7.0% 와 맞먹는다.**
#    EUA 는 2020 €25 → 2023 €100 을 오갔다. €30 움직이면 33 ₩/kWh(17%)다.
#
#  ■ 지금 위험의 성격 — 정확도가 아니라 '못 보는 것'
#    폴란드 도매가에는 EUA 가 **이미 들어 있다**(한계발전기 탄소비가 시장가에 실린다).
#    그래서 수준은 틀리지 않았다. 빠진 것은 **미래 EUA 를 보고 미리 아는 경로**다.
#    지금은 도매가 실적 추세만 이어서, EUA 가 뛰면 도매가가 뛴 뒤에야 따라간다 —
#    **오하이오와 같은 실패 형태다.** 폴란드 MAPE 7.0% 는 최근 44개월이 조용했기 때문이다.
#
#  ■ 1차 실행에서 EUA 만 실패한 이유와 대응
#    Yahoo 429(러너 공용 IP 요율제한) · Stooq 봇차단(JS 챌린지) · EEX gvsi 연결거부.
#    반면 **EEX 경매 페이지와 ICAP 페이지는 200(HTML)** 이었다. 그래서 여기서는
#      (1) 그 페이지 본문에서 **데이터파일 링크·API 경로를 발견**한다(World Bank 방식)
#      (2) Yahoo 는 백오프 + query2 호스트로 재시도한다(KEUA·KRBN 은 EUA 선물 보유 ETF)
#      (3) FastAPI 계열은 **openapi.json 을 받아 엔드포인트 목록을 읽는다** — 경로를
#          상상하지 않고 서버가 스스로 알려주게 한다
# ══════════════════════════════════════════════════════════════════
DATE_RE = re.compile(r"^\s*(\d{4})[-/.](\d{1,2})[-/.](\d{1,2})")
EP_RE   = re.compile(r"""["'](https?://[^"'\s]{12,200}|/[A-Za-z0-9._/\-]{6,140})["']""")
EP_KEY  = re.compile(r"(api|\.json|\.csv|\.xlsx?|graphql|price|auction|chart|data)", re.I)
FILE2_RE= re.compile(r"""["'(]([^"'()\s]{4,240}\.(?:csv|xlsx|xls|json))["')]""", re.I)
PRICE_H = re.compile(r"(price|settle|clos|eur|auction|value)", re.I)
MIN_DAYS = 15          # 월평균은 관측 15일 이상인 달만 인정(부분월 편향 방지)


def _monthly_from_daily(pairs):
    """pairs=[(YYYY-MM-DD, float)] → [(YYYY-MM, 월평균)] · 관측 15일 미만 달은 버린다."""
    acc = {}
    for d, v in pairs:
        acc.setdefault(d[:7], []).append(v)
    out = []
    for ym in sorted(acc):
        if len(acc[ym]) >= MIN_DAYS:
            out.append((ym, sum(acc[ym]) / len(acc[ym])))
    return out


def parse_csv_prices(text):
    """헤더에서 날짜열·가격열을 찾아 (일별 또는 월별) 시계열을 만든다.
       열 이름을 고정하지 않는다 — 날짜 매치가 가장 많은 열을 날짜열로 본다."""
    try:
        rows = list(csv.reader(io.StringIO(text)))
    except Exception:
        return [], "csv 파싱 실패"
    if len(rows) < 8:
        return [], "행 %d개 (너무 적음)" % len(rows)
    # 날짜열 = DATE_RE 매치가 가장 많은 열
    ncol = max(len(r) for r in rows[:200])
    hit = [0] * ncol
    for r in rows:
        for c in range(min(ncol, len(r))):
            if DATE_RE.match(str(r[c])):
                hit[c] += 1
    if not any(hit):
        return [], "날짜열 없음 (헤더=%s)" % (rows[0][:6],)
    dcol = hit.index(max(hit))
    hdr = rows[0]
    # 가격열 = 헤더가 price/settle/close/eur 계열이고 숫자가 채워진 열
    cands = [c for c, h in enumerate(hdr) if c != dcol and PRICE_H.search(str(h))]
    if not cands:
        cands = [c for c in range(ncol) if c != dcol]
    best, bn = None, 0
    for c in cands:
        n = 0
        for r in rows[1:]:
            if len(r) > max(c, dcol) and DATE_RE.match(str(r[dcol])):
                try:
                    float(str(r[c]).replace(",", "").replace(" ", ""))
                    n += 1
                except (TypeError, ValueError):
                    pass
        if n > bn:
            best, bn = c, n
    if best is None or bn < 8:
        return [], "가격열 없음 (헤더=%s)" % (hdr[:8],)
    pairs = []
    for r in rows[1:]:
        if len(r) <= max(best, dcol):
            continue
        m = DATE_RE.match(str(r[dcol]))
        if not m:
            continue
        d = "%04d-%02d-%02d" % (int(m.group(1)), int(m.group(2)), int(m.group(3)))
        try:
            v = float(str(r[best]).replace(",", "").replace(" ", ""))
        except (TypeError, ValueError):
            continue
        if v > 0:
            pairs.append((d, v))
    return _monthly_from_daily(pairs), "날짜열%d 가격열%d(%s) 일별 %d건" % (
        dcol, best, str(hdr[best])[:24], len(pairs))


def xlsx_dates(grid, price_rx=PRICE_H):
    """날짜 셀(datetime) 또는 'YYYY-MM-DD' 문자열을 쓰는 xlsx 용 파서.
       World Bank·IMF 는 '1960M01' 형식이라 xlsx_pick 이 맞지만, 거래소·경매 보고서는
       **실제 날짜 셀**을 쓴다(openpyxl 이 datetime 객체로 준다). 두 경로를 나눠 둔다."""
    import datetime as _dt

    def asdate(v):
        if isinstance(v, (_dt.datetime, _dt.date)):
            return "%04d-%02d-%02d" % (v.year, v.month, v.day)
        m = DATE_RE.match(str(v)) if v is not None else None
        return "%04d-%02d-%02d" % (int(m.group(1)), int(m.group(2)), int(m.group(3))) if m else None

    ncol = max((len(r) for r in grid), default=0)
    if not ncol:
        return [], "빈 시트"
    hit = [0] * ncol
    for r in grid:
        for c in range(min(ncol, len(r))):
            if asdate(r[c]):
                hit[c] += 1
    if not any(hit):
        return [], "날짜 셀 없음"
    dcol = hit.index(max(hit))
    first = next(i for i, r in enumerate(grid) if len(r) > dcol and asdate(r[dcol]))
    # 헤더 = 첫 데이터행 위쪽을 열 단위로 이어붙임(xlsx_pick 과 같은 규칙)
    hdr = []
    for c in range(ncol):
        buf = [str(grid[i][c]).strip() for i in range(max(0, first - 12), first)
               if len(grid[i]) > c and isinstance(grid[i][c], str) and grid[i][c].strip()]
        hdr.append(" ".join(buf))
    cands = [c for c in range(ncol) if c != dcol and hdr[c] and price_rx.search(hdr[c])] \
            or [c for c in range(ncol) if c != dcol]
    best, bn = None, 0
    for c in cands:
        n = sum(1 for r in grid[first:]
                if len(r) > max(c, dcol) and asdate(r[dcol]) and isinstance(r[c], (int, float)))
        if n > bn:
            best, bn = c, n
    if best is None or bn < 8:
        return [], "가격열 없음(헤더=%s)" % [h[:20] for h in hdr[:8]]
    pairs = []
    for r in grid[first:]:
        if len(r) <= max(best, dcol):
            continue
        d = asdate(r[dcol])
        if d and isinstance(r[best], (int, float)) and r[best] > 0:
            pairs.append((d, float(r[best])))
    return _monthly_from_daily(pairs), "날짜열%d 가격열%d(%s) 일별 %d건" % (
        dcol, best, hdr[best][:24], len(pairs))


def page_hunt(pid, url, key_rx):
    """페이지를 받아 (a) 데이터파일 링크 (b) API 경로 후보를 **발견**한다.
       링크가 csv/xlsx 면 실제로 내려받아 파싱까지 시도한다. 못 하면 프로브로만 남긴다."""
    try:
        r = get(url)
    except Exception as e:
        rec(pid, 1, url, None, False, "EXC %s" % e)
        return 0
    if r.status_code != 200:
        rec(pid, 1, url, r.status_code, False, r.text[:150])
        return 0
    body = r.text
    files, eps = [], []
    for f in FILE2_RE.findall(body):
        if key_rx.search(f):
            files.append(f if f.startswith("http") else _abs(url, f))
    for e in EP_RE.findall(body):
        if EP_KEY.search(e) and key_rx.search(e):
            eps.append(e if e.startswith("http") else _abs(url, e))
    files = list(dict.fromkeys(files))[:12]
    eps = list(dict.fromkeys(eps))[:15]
    rec(pid, 1, url, 200, bool(files or eps),
        "파일 %d건 %s | 경로후보 %d건 %s" % (len(files), files[:4], len(eps), eps[:4]))

    n = 0
    for fi, f in enumerate(files):        # 한 페이지가 파일 여러 개를 주면 키가 겹쳐 덮어쓴다 -> 인덱스 부여
        tag = pid + (("_%d" % (fi + 1)) if fi else "")
        try:
            rr = get(f)
            if rr.status_code != 200:
                rec(tag + "_dl", 1, f, rr.status_code, False, "")
                continue
            if f.lower().endswith((".xlsx", ".xls")):
                sheet, grid = xlsx_grid(rr.content, "")
                got, hrow, dcol = xlsx_pick(grid, {"EUA": (r"eua|allowance|auction|settle|price",
                                                           "EUA (xlsx)")})
                if got.get("EUA"):
                    k = put("EUA_" + tag, "EUA 배출권 (xlsx: " + f.rsplit("/", 1)[-1][:32] + ")",
                            "EUR/tCO2", url, got["EUA"])
                    n += k
                    rec(tag + "_dl", 1, f, 200, True, "xlsx(YYYYMnn) sheet=%s pts=%d" % (sheet, k))
                else:
                    # 거래소·경매 보고서는 실제 날짜 셀을 쓴다 → 두 번째 파서로 재시도
                    pairs, note2 = xlsx_dates(grid)
                    k = put("EUA_" + tag, "EUA 배출권 (xlsx: " + f.rsplit("/", 1)[-1][:32] + ")",
                            "원문단위 · 확인 필요", url, pairs) if pairs else 0
                    n += k
                    rec(tag + "_dl", 1, f, 200, k > 0,
                        "xlsx(YYYYMnn) 미발견 hdr=%s dcol=%s → 날짜셀 파서: %s → 월 %d건"
                        % (hrow, dcol, note2, len(pairs)))
            else:
                pairs, note = parse_csv_prices(rr.text)
                if pairs:
                    n += put("EUA_" + tag, "EUA 배출권 (csv: " + f.rsplit("/", 1)[-1][:32] + ")",
                             "EUR/tCO2(추정 · 단위 확인 필요)", url, pairs)
                rec(tag + "_dl", 1, f, 200, bool(pairs), note + " → 월 %d건" % len(pairs))
        except Exception as e:
            rec(tag + "_dl", 1, f, None, False, "EXC %s" % e)
    return n


def _abs(base, href):
    if href.startswith("http"):
        return href
    m = re.match(r"(https?://[^/]+)", base)
    root = m.group(1) if m else ""
    return (root + href) if href.startswith("/") else (base.rsplit("/", 1)[0] + "/" + href)


EUA_KEY = re.compile(r"(eua|emission|allowance|carbon|co2|auction|ets)", re.I)
EUA_PAGES = [
    ("EEXauc",  "https://www.eex.com/en/market-data/environmental-markets/eua-primary-auction-spot-download"),
    ("EEXemis", "https://www.eex.com/en/market-data/environmental-markets"),
    ("ICAP",    "https://icapcarbonaction.com/en/ets-prices"),
]
# FastAPI 계열은 스펙을 스스로 알려준다 — 경로를 상상하지 않는다.
OPENAPI = [
    ("EC_spec",    "https://api.energy-charts.info/openapi.json"),
    ("EMBER_spec", "https://api.ember-energy.org/openapi.json"),
]
YH_EUA = ["KEUA", "KRBN"]          # EUA 선물을 실제 보유하는 ETF (대리지표)


def src_eua():
    n = 0
    for pid, u in EUA_PAGES:
        try:
            n += page_hunt("EUA_" + pid, u, EUA_KEY)
        except Exception as e:
            rec("EUA_" + pid, 1, u, None, False, "EXC %s" % e)

    # openapi 스펙 → carbon/co2/price 관련 경로 목록만 뽑아 로그에 남긴다(호출은 안 한다)
    for pid, u in OPENAPI:
        try:
            r = get(u)
            if r.status_code != 200:
                rec(pid, 1, u, r.status_code, False, r.text[:120])
                continue
            j = r.json()
            paths = list((j.get("paths") or {}).keys())
            hitp = [p for p in paths if re.search(r"co2|carbon|price|emission", p, re.I)]
            rec(pid, 1, u, 200, bool(hitp),
                "전체 %d경로 · 관련 %s" % (len(paths), hitp[:10]))
        except Exception as e:
            rec(pid, 1, u, None, False, "EXC %s" % e)

    # Yahoo 재시도 — 1차 실행에서 5티커 전부 429 였다. 호스트 교대 + 백오프.
    for sym in YH_EUA:
        done = False
        for attempt, host in enumerate(["query1", "query2", "query1"]):
            u = ("https://%s.finance.yahoo.com/v8/finance/chart/%s"
                 "?range=10y&interval=1mo" % (host, sym))
            try:
                if attempt:
                    time.sleep(4 * attempt)          # 4s · 8s 백오프
                r = get(u)
                if r.status_code != 200:
                    rec("YHR_%s#%d" % (sym, attempt + 1), 1, u, r.status_code, False,
                        r.text[:90])
                    continue
                j = r.json()
                res = ((j.get("chart") or {}).get("result") or [None])[0]
                if not res:
                    rec("YHR_%s#%d" % (sym, attempt + 1), 1, u, 200, False, "result 없음")
                    continue
                ts = res.get("timestamp") or []
                ind = res.get("indicators") or {}
                vals = ((ind.get("adjclose") or [{}])[0].get("adjclose")
                        or (ind.get("quote") or [{}])[0].get("close") or [])
                pairs = []
                for t, v in zip(ts, vals):
                    if v is None:
                        continue
                    g = time.gmtime(t)
                    pairs.append(("%04d-%02d" % (g.tm_year, g.tm_mon), float(v)))
                if pairs:
                    pairs = pairs[:-1]               # ★ 마지막은 진행 중인 달 — 월평균 아님
                k = put("YH_" + sym, sym + " (EUA 선물 보유 ETF · 대리지표)", "USD/주",
                        "Yahoo Finance chart v8 (" + host + ")", pairs)
                n += k
                rec("YHR_%s#%d" % (sym, attempt + 1), 1, u, 200, k > 0,
                    "pts=%d %s~%s" % (k, pairs[0][0] if pairs else "-",
                                      pairs[-1][0] if pairs else "-"))
                done = True
                break
            except Exception as e:
                rec("YHR_%s#%d" % (sym, attempt + 1), 1, u, None, False, "EXC %s" % e)
        if not done:
            log("   !! %s 3회 모두 실패 — 러너 IP 요율제한이 지속적일 수 있다" % sym)
    return n


# ══════════════════════════════════════════════════════════════════
#  T3  forward 곡선·1차시장 후보 — 파서 없음. 살아있는지만 본다.
#      여기서 200 이 나오면 FUEL_FWD 자동화가 열린다(payoff 최대).
#      쿼리 파라미터를 추측해 채우지 않는다 — 200 을 받은 뒤 본문 보고 짠다.
# ══════════════════════════════════════════════════════════════════
FWD_PROBE = [
    ("EEX_gvsi",  "https://webservice-eex.gvsi.com/query/json/getDaily/close/tradedatetimegmt/?symbol=%2FE.FTBM"),
    ("EEX_group", "https://api.eex-group.com/"),
    ("EEX_auc",   "https://www.eex.com/en/market-data/environmental-markets/eua-primary-auction-spot-download"),
    ("EEX_gas",   "https://www.eex.com/en/market-data/natural-gas/futures"),
    ("ICE_report","https://www.theice.com/marketdata/reports/api/getReport?reportId=178"),
    ("ICAP_api",  "https://icapcarbonaction.com/api/allowance-prices"),
    ("ICAP_page", "https://icapcarbonaction.com/en/ets-prices"),
    ("EMBER_api", "https://api.ember-energy.org/v1/carbon-price/monthly"),
    ("EUROSTAT",  "https://ec.europa.eu/eurostat/api/dissemination/statistics/1.0/data/nrg_pc_203"
                  "?format=JSON&lang=EN&geo=PL"),
]


def src_probe_only():
    for pid, u in FWD_PROBE:
        try:
            r = get(u, allow_redirects=True)
            ct = r.headers.get("Content-Type", "")
            if "text" in ct or "json" in ct or "xml" in ct:
                head = re.sub(r"\s+", " ", r.text[:220])
            else:
                head = "<bin %dB>" % len(r.content)
            rec(pid, 3, u, r.status_code, r.status_code == 200, "ct=%s | %s" % (ct, head))
        except Exception as e:
            rec(pid, 3, u, None, False, "EXC %s" % e)


# ══════════════════════════════════════════════════════════════════
def main():
    log("=== fetch_fuel.py  mode=%s ===" % ("DISCOVER" if DISCOVER else "COLLECT"))
    for fn in (src_worldbank, src_imf, src_yahoo, src_stooq, src_eua, src_probe_only):
        try:
            fn()
        except Exception as e:
            log("!! %s 전체 실패: %s" % (fn.__name__, e))

    ok = [p for p in PROBE if p["ok"]]
    log("\n[요약] 후보 %d건 중 성공 %d건 · 계열 %d개" % (len(PROBE), len(ok), len(SERIES)))
    for k, s in SERIES.items():
        log("  %-22s n=%4d %s~%s  %s" % (k, len(s["points"]),
            s["points"][0]["period"], s["points"][-1]["period"], s["label"]))

    with open(OUT_PROBE, "w", encoding="utf-8") as f:
        json.dump({"probe": PROBE, "seriesFound": list(SERIES)}, f, ensure_ascii=False, indent=2)
    log("[OK] %s" % OUT_PROBE)

    if DISCOVER:
        log("[SKIP] --discover 이므로 fuel_monthly.* 는 갱신하지 않는다")
        return
    if not SERIES:
        log("!! 계열 0개 — 출력파일을 덮어쓰지 않는다(기존 수집분 보호)")
        return

    with open(OUT_JSON, "w", encoding="utf-8") as f:
        json.dump({"source": "무료 공개소스 다중(World Bank·IMF·Yahoo·Stooq)",
                   "note": "월별 현물/월평균이다. forward 아님. 단위는 원문 그대로 — 환산은 파일럿에서.",
                   "data": SERIES}, f, ensure_ascii=False, indent=2)
    log("[OK] %s" % OUT_JSON)

    with open(OUT_CSV, "w", newline="", encoding="utf-8-sig") as f:
        w = csv.writer(f)
        w.writerow(["series", "period", "value", "unit", "label", "source"])
        for k, s in SERIES.items():
            for p in s["points"]:
                w.writerow([k, p["period"], p["value"], s["unit"], s["label"], s["source"]])
    log("[OK] %s" % OUT_CSV)


if __name__ == "__main__":
    main()
